"""
Excel Structure Detection Agent
Analyzes Excel files to detect table structure, headers, and data organization
"""

import pandas as pd
import openpyxl
from typing import Dict, Any, List, Tuple, Optional
from pathlib import Path
import re

from agent_base import BaseAgent, AgentResponse
from logger import ProcessingLogger


class ExcelStructureAgent(BaseAgent):
    """Agent specialized in detecting Excel table structure and organization."""

    def __init__(self, logger: ProcessingLogger, api_client=None):
        super().__init__("excel_structure_agent", logger, api_client)
        self.current_analysis = {}

    def get_system_prompt(self) -> str:
        """Get the system prompt for structure detection."""
        return """You are an Excel structure analysis expert. Your role is to:
1. Detect table layouts and organization patterns
2. Identify header rows and their levels
3. Determine category/subcategory columns
4. Identify data columns and their types
5. Detect merged cells and their implications
6. Understand hierarchical data relationships"""

    def process(self, input_data: Any, context: Dict[str, Any] = None) -> AgentResponse:
        """Analyze Excel structure and return detailed structure information."""
        try:
            if isinstance(input_data, dict) and "worksheet" in input_data:
                # Direct worksheet analysis
                worksheet = input_data["worksheet"]
                sheet_name = input_data.get("sheet_name", "Sheet")
                structure = self._analyze_worksheet_structure(worksheet)
            elif isinstance(input_data, dict) and "dataframe" in input_data:
                # Pandas DataFrame analysis
                df = input_data["dataframe"]
                sheet_name = input_data.get("sheet_name", "Sheet")
                structure = self._analyze_dataframe_structure(df)
            else:
                raise ValueError("Input must contain either 'worksheet' or 'dataframe'")

            # Store analysis for reference
            self.current_analysis = structure

            return AgentResponse(
                success=True,
                content=structure,
                confidence=self._calculate_confidence(structure),
                metadata={"sheet_name": sheet_name}
            )

        except Exception as e:
            self.logger.log_error(f"Structure detection failed: {str(e)}")
            return AgentResponse(
                success=False,
                content={},
                confidence=0.0,
                error_message=str(e)
            )

    def _analyze_worksheet_structure(self, worksheet) -> Dict[str, Any]:
        """Analyze openpyxl worksheet structure."""
        structure = {
            "type": "openpyxl_worksheet",
            "dimensions": {
                "rows": worksheet.max_row,
                "columns": worksheet.max_column
            },
            "merged_cells": [],
            "header_info": {},
            "data_start_row": 0,
            "column_structure": [],
            "hierarchical": False
        }

        # Detect merged cells
        merged_map = {}
        for merged_range in worksheet.merged_cells.ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            max_row, max_col = merged_range.max_row, merged_range.max_col

            value = worksheet.cell(min_row, min_col).value

            structure["merged_cells"].append({
                "range": f"{min_row}:{min_col}-{max_row}:{max_col}",
                "value": value,
                "spans_columns": max_col - min_col + 1,
                "spans_rows": max_row - min_row + 1
            })

            # Map for later use
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged_map[(row, col)] = value

        # Read sample data for analysis
        sample_rows = []
        for row_idx in range(1, min(20, worksheet.max_row + 1)):
            row_data = []
            for col_idx in range(1, worksheet.max_column + 1):
                if (row_idx, col_idx) in merged_map:
                    value = merged_map[(row_idx, col_idx)]
                else:
                    value = worksheet.cell(row_idx, col_idx).value
                row_data.append(value)
            sample_rows.append(row_data)

        # Analyze structure
        header_info = self._detect_headers(sample_rows)
        structure["header_info"] = header_info
        structure["data_start_row"] = header_info.get("data_start_row", 1)

        # Analyze column structure
        structure["column_structure"] = self._analyze_columns(
            sample_rows,
            header_info.get("header_rows", [])
        )

        # Detect if hierarchical
        structure["hierarchical"] = self._detect_hierarchy(sample_rows, structure["data_start_row"])

        return structure

    def _analyze_dataframe_structure(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Analyze pandas DataFrame structure."""
        structure = {
            "type": "pandas_dataframe",
            "dimensions": {
                "rows": len(df),
                "columns": len(df.columns)
            },
            "merged_cells": [],  # Can't detect from DataFrame
            "header_info": {},
            "data_start_row": 0,
            "column_structure": [],
            "hierarchical": False
        }

        # Convert DataFrame to list for analysis
        sample_rows = []

        # Add column names as potential first header row
        sample_rows.append(list(df.columns))

        # Add data rows
        for idx, row in df.head(20).iterrows():
            sample_rows.append(row.tolist())

        # Analyze structure
        header_info = self._detect_headers(sample_rows)
        structure["header_info"] = header_info

        # Analyze columns
        structure["column_structure"] = self._analyze_dataframe_columns(df)

        # Detect hierarchy
        structure["hierarchical"] = self._detect_dataframe_hierarchy(df)

        return structure

    def _detect_headers(self, rows: List[List]) -> Dict[str, Any]:
        """Detect header rows in the data."""
        header_info = {
            "header_rows": [],
            "header_levels": 0,
            "data_start_row": len(rows),
            "has_multi_level": False
        }

        # Be conservative - only check first few rows and stop at first data row
        for idx, row in enumerate(rows[:3]):  # Only check first 3 rows max
            if not row or all(v is None for v in row):
                continue

            # Count non-empty cells
            non_empty = sum(1 for v in row if v is not None and str(v).strip())
            if non_empty == 0:
                continue

            # First row is almost always a header if it has content
            if idx == 0:
                header_info["header_rows"].append(idx)
                continue

            # For subsequent rows, be very strict about what constitutes a header
            # Only consider it a header if it contains mostly descriptive text (not ratios, percentages, etc.)
            likely_data_values = sum(1 for v in row if v is not None and
                                   (self._is_numeric(str(v)) or
                                    ':' in str(v) or  # ratios like 1.2:1
                                    '%' in str(v) or  # percentages
                                    str(v).strip().lower() in ['yes', 'no']))  # simple yes/no values

            # If more than 25% of values look like data rather than headers, treat as data row
            if likely_data_values / non_empty > 0.25:
                header_info["data_start_row"] = idx
                break
            else:
                header_info["header_rows"].append(idx)

        # If we didn't find a clear data start, assume it starts after headers
        if header_info["data_start_row"] == len(rows):
            header_info["data_start_row"] = len(header_info["header_rows"])

        header_info["header_levels"] = len(header_info["header_rows"])
        header_info["has_multi_level"] = header_info["header_levels"] > 1

        return header_info

    def _analyze_columns(self, rows: List[List], header_indices: List[int]) -> List[Dict]:
        """Analyze column structure and types."""
        if not rows:
            return []

        num_cols = max(len(row) for row in rows)
        columns = []

        for col_idx in range(num_cols):
            col_info = {
                "index": col_idx,
                "header": self._build_column_header(rows, header_indices, col_idx),
                "type": "unknown",
                "role": "data",  # category, subcategory, or data
                "sample_values": []
            }

            # Collect sample values (skip headers)
            data_start = max(header_indices) + 1 if header_indices else 0
            for row_idx in range(data_start, min(data_start + 5, len(rows))):
                if row_idx < len(rows) and col_idx < len(rows[row_idx]):
                    value = rows[row_idx][col_idx]
                    if value is not None and str(value).strip():
                        col_info["sample_values"].append(str(value))

            # Determine column type and role
            col_info["type"] = self._detect_column_type(col_info["sample_values"])
            col_info["role"] = self._detect_column_role(col_idx, col_info)

            columns.append(col_info)

        return columns

    def _analyze_dataframe_columns(self, df: pd.DataFrame) -> List[Dict]:
        """Analyze DataFrame column structure."""
        columns = []

        for col_idx, col_name in enumerate(df.columns):
            col_info = {
                "index": col_idx,
                "header": str(col_name),
                "type": str(df[col_name].dtype),
                "role": "data",
                "sample_values": []
            }

            # Get sample values
            sample = df[col_name].dropna().head(5)
            col_info["sample_values"] = [str(v) for v in sample]

            # Detect role
            if col_idx == 0:
                # First column often category
                non_empty_ratio = df[col_name].notna().sum() / len(df)
                if non_empty_ratio > 0.3:  # At least 30% filled
                    col_info["role"] = "category"
            elif col_idx == 1 and len(df.columns) > 2:
                # Second column might be subcategory
                col_info["role"] = "subcategory"

            columns.append(col_info)

        return columns

    def _detect_hierarchy(self, rows: List[List], data_start: int) -> bool:
        """Detect if data has hierarchical structure."""
        if not rows or data_start >= len(rows):
            return False

        # Check first column for sparse data (indicates category column)
        first_col_values = []
        for row in rows[data_start:min(data_start + 10, len(rows))]:
            if row and len(row) > 0:
                first_col_values.append(row[0])

        # Count non-None values
        non_none = sum(1 for v in first_col_values if v is not None and str(v).strip())

        # If first column is sparse (less than 70% filled), likely hierarchical
        if len(first_col_values) > 0:
            fill_ratio = non_none / len(first_col_values)
            return fill_ratio < 0.7

        return False

    def _detect_dataframe_hierarchy(self, df: pd.DataFrame) -> bool:
        """Detect hierarchy in DataFrame."""
        if len(df.columns) < 2:
            return False

        # Check if first column has many repeated or empty values
        first_col = df.iloc[:, 0]
        unique_ratio = first_col.nunique() / len(df)

        # Low unique ratio suggests hierarchical structure
        return unique_ratio < 0.5

    def _build_column_header(self, rows: List[List], header_indices: List[int], col_idx: int) -> str:
        """Build column header from potentially multiple header rows."""
        if not header_indices:
            return f"Column{col_idx + 1}"

        header_parts = []
        for header_idx in header_indices:
            if header_idx < len(rows) and col_idx < len(rows[header_idx]):
                value = rows[header_idx][col_idx]
                if value is not None and str(value).strip():
                    value_str = str(value).strip()
                    if not header_parts or header_parts[-1] != value_str:
                        header_parts.append(value_str)

        if header_parts:
            return " - ".join(header_parts)
        return f"Column{col_idx + 1}"

    def _detect_column_type(self, sample_values: List[str]) -> str:
        """Detect the type of data in a column."""
        if not sample_values:
            return "empty"

        # Check for numeric
        numeric_count = sum(1 for v in sample_values if self._is_numeric(v))
        if numeric_count == len(sample_values):
            return "numeric"

        # Check for currency
        currency_count = sum(1 for v in sample_values if '$' in v or self._is_numeric(v.replace('$', '').replace(',', '')))
        if currency_count == len(sample_values):
            return "currency"

        # Check for percentage
        percent_count = sum(1 for v in sample_values if '%' in v or (self._is_numeric(v.replace('%', '')) and float(v.replace('%', '')) <= 100))
        if percent_count == len(sample_values):
            return "percentage"

        # Check for dates
        date_count = sum(1 for v in sample_values if self._looks_like_date(v))
        if date_count > len(sample_values) * 0.5:
            return "date"

        return "text"

    def _detect_column_role(self, col_idx: int, col_info: Dict) -> str:
        """Determine the role of a column in the data structure."""
        # Text columns at the beginning are likely label columns
        if col_info["type"] == "text":
            # Check if this is a label column by examining sample values
            samples = col_info.get("sample_values", [])
            if samples and all(len(s) < 100 for s in samples):  # Reasonable label length
                # Assign specific label role based on position
                if col_idx == 0:
                    return "label_1"
                elif col_idx == 1:
                    return "label_2"
                elif col_idx == 2:
                    return "label_3"
                elif col_idx == 3:
                    return "label_4"
                elif col_idx == 4:
                    return "label_5"

        # Numeric columns are usually data
        if col_info["type"] in ["numeric", "currency", "percentage"]:
            return "data"

        return "data"

    def _is_numeric(self, value: str) -> bool:
        """Check if value is numeric."""
        if not value:
            return False
        value = str(value).strip()
        # Remove common formatting
        test_value = value.replace(',', '').replace('$', '').replace('%', '').replace('(', '').replace(')', '')
        try:
            float(test_value)
            return True
        except (ValueError, TypeError):
            return False

    def _looks_like_date(self, value: str) -> bool:
        """Check if value looks like a date."""
        date_patterns = [
            r'^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$',
            r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}$',
            r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)',
            r'^\d{1,2}-(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)',
        ]
        value = str(value).strip()
        return any(re.match(pattern, value, re.IGNORECASE) for pattern in date_patterns)

    def _calculate_confidence(self, structure: Dict) -> float:
        """Calculate confidence in structure detection."""
        confidence = 0.5

        # Boost for clear headers
        if structure.get("header_info", {}).get("header_rows"):
            confidence += 0.2

        # Boost for consistent column types
        columns = structure.get("column_structure", [])
        if columns:
            typed_cols = sum(1 for c in columns if c["type"] != "unknown")
            confidence += (typed_cols / len(columns)) * 0.2

        # Boost for hierarchy detection
        if structure.get("hierarchical") is not None:
            confidence += 0.1

        return min(1.0, confidence)