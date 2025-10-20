"""
Excel Formatting Agent
Formats Excel data based on detected structure into structured markdown lists
"""

import pandas as pd
import openpyxl
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime

from agent_base import BaseAgent, AgentResponse
from logger import ProcessingLogger


class ExcelFormattingAgent(BaseAgent):
    """Agent specialized in formatting Excel data based on detected structure."""

    def __init__(self, logger: ProcessingLogger, api_client=None):
        super().__init__("excel_formatting_agent", logger, api_client)

    def get_system_prompt(self) -> str:
        """Get the system prompt for formatting."""
        return """You are an Excel data formatting expert. Your role is to:
1. Format data based on detected table structure
2. Preserve all context relationships for vector storage
3. Create consistent hierarchical output patterns
4. Ensure each data point maintains full context
5. Handle merged cells and multi-level headers properly"""

    def process(self, input_data: Any, context: Dict[str, Any] = None) -> AgentResponse:
        """Format Excel data based on structure analysis."""
        try:
            # Extract required components
            structure = input_data.get("structure", {})
            data_source = input_data.get("data_source")
            document_name = input_data.get("document_name", "Document")
            sheet_name = input_data.get("sheet_name", "Sheet")

            if not structure:
                raise ValueError("Structure analysis required for formatting")

            # Format based on data source type
            if structure.get("type") == "openpyxl_worksheet":
                formatted_content = self._format_openpyxl_data(
                    data_source, structure, document_name, sheet_name
                )
            elif structure.get("type") == "pandas_dataframe":
                formatted_content = self._format_dataframe_data(
                    data_source, structure, document_name, sheet_name
                )
            else:
                raise ValueError(f"Unsupported data source type: {structure.get('type')}")

            return AgentResponse(
                success=True,
                content=formatted_content,
                confidence=self._calculate_formatting_confidence(formatted_content, structure),
                metadata={
                    "document_name": document_name,
                    "sheet_name": sheet_name,
                    "structure_type": structure.get("type"),
                    "hierarchical": structure.get("hierarchical", False)
                }
            )

        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            self.logger.log_error(f"Formatting failed: {str(e)}\n{error_details}")
            return AgentResponse(
                success=False,
                content="",
                confidence=0.0,
                error_message=f"{str(e)}\n{error_details}"
            )

    def _format_openpyxl_data(self, worksheet, structure: Dict, doc_name: str, sheet_name: str) -> str:
        """Format openpyxl worksheet data with structure awareness."""
        # Skip document header - just start with data
        content = []

        # Build merged cells map
        merged_map = self._build_merged_cells_map(worksheet, structure.get("merged_cells", []))

        # Extract all data
        all_rows = self._extract_worksheet_data(worksheet, merged_map)

        if not all_rows:
            content.append("- *[No data in this sheet]*")
            return "\n".join(content)

        # Get structure info
        data_start = structure.get("data_start_row", 0)
        columns = structure.get("column_structure", [])
        is_hierarchical = structure.get("hierarchical", False)
        include_headers = structure.get("include_headers", False)

        # Extract data rows
        data_rows = all_rows[data_start:] if data_start < len(all_rows) else []

        # Format the data
        formatted_lines = self._format_rows_with_structure(
            data_rows, columns, is_hierarchical, include_headers
        )

        content.extend(formatted_lines)
        return "\n".join(content)

    def _format_dataframe_data(self, df: pd.DataFrame, structure: Dict, doc_name: str, sheet_name: str) -> str:
        """Format pandas DataFrame with structure awareness."""
        # Skip document header - just start with data
        content = []

        if df.empty:
            content.append("- *[No data in this sheet]*")
            return "\n".join(content)

        # Get structure info
        columns = structure.get("column_structure", [])
        is_hierarchical = structure.get("hierarchical", False)
        include_headers = structure.get("include_headers", False)

        # Convert DataFrame to rows
        data_rows = []
        for _, row in df.iterrows():
            data_rows.append(row.tolist())

        # Format the data
        formatted_lines = self._format_rows_with_structure(
            data_rows, columns, is_hierarchical, include_headers
        )

        content.extend(formatted_lines)
        return "\n".join(content)

    def _build_document_header(self, doc_name: str, sheet_count: int) -> List[str]:
        """Build standard document header."""
        return [
            f"# {doc_name}",
            "",
            f"**Processed:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"**Source:** Excel file with {sheet_count} sheet(s)",
            "",
            "---",
            ""
        ]

    def _build_merged_cells_map(self, worksheet, merged_cells_info: List[Dict]) -> Dict:
        """Build map of merged cells from structure info."""
        merged_map = {}

        for merged_info in merged_cells_info:
            # Parse range string like "1:2-3:4"
            range_str = merged_info.get("range", "")
            value = merged_info.get("value")

            if ":" in range_str and "-" in range_str:
                start_part, end_part = range_str.split("-")
                start_row, start_col = map(int, start_part.split(":"))
                end_row, end_col = map(int, end_part.split(":"))

                for row in range(start_row, end_row + 1):
                    for col in range(start_col, end_col + 1):
                        merged_map[(row, col)] = value

        return merged_map

    def _extract_worksheet_data(self, worksheet, merged_map: Dict) -> List[List]:
        """Extract all data from worksheet including merged cells."""
        data_rows = []

        for row_idx in range(1, worksheet.max_row + 1):
            row_data = []
            for col_idx in range(1, worksheet.max_column + 1):
                if (row_idx, col_idx) in merged_map:
                    value = merged_map[(row_idx, col_idx)]
                else:
                    value = worksheet.cell(row_idx, col_idx).value
                row_data.append(value)
            data_rows.append(row_data)

        return data_rows

    def _format_rows_with_structure(self, data_rows: List[List], columns: List[Dict], is_hierarchical: bool, include_headers: bool = False) -> List[str]:
        """Format data rows based on detected structure."""
        if not data_rows:
            return ["- *[No data]*"]

        content = []

        # Determine column roles - collect all label columns
        label_cols = [c for c in columns if c.get("role", "").startswith("label_")]
        # Sort label columns by their index to ensure correct order
        label_cols.sort(key=lambda x: x["index"])

        # Legacy support: fall back to category/subcategory if no label columns found
        if not label_cols:
            category_col = self._find_column_by_role(columns, "category")
            subcategory_col = self._find_column_by_role(columns, "subcategory")
            if category_col:
                label_cols.append(category_col)
            if subcategory_col:
                label_cols.append(subcategory_col)

        data_cols = [c for c in columns if c.get("role") == "data"]

        # First pass: collect all formatted lines with their grouping info
        all_lines = []
        for row in data_rows:
            if not row or all(v is None or str(v).strip() == '' for v in row):
                continue

            # Format row with new label concatenation logic
            formatted_lines = self._format_row_with_labels(row, label_cols, data_cols)

            # Add lines with grouping context (only data lines, not headers)
            for line in formatted_lines:
                if line.startswith("- "):  # Only count actual data lines
                    group_key = self._extract_group_key(line)
                    all_lines.append((line, group_key))

        # Second pass: add headers if requested
        if include_headers:
            content = self._add_section_headers(all_lines)
        else:
            # Just return the lines without headers
            content = [line for line, _ in all_lines]

        return content if content else ["- *[No readable data]*"]

    def _format_row_with_labels(self, row: List, label_cols: List[Dict], data_cols: List[Dict]) -> List[str]:
        """Format a row by concatenating all label columns into a single identifier."""
        lines = []

        # Build row identifier from all label columns
        label_parts = []
        for label_col in label_cols:
            if label_col["index"] < len(row):
                value = row[label_col["index"]]
                if value is not None and str(value).strip():
                    label_parts.append(str(value).strip())

        if not label_parts:
            # No labels found, skip this row
            return lines

        # Join all label parts with " - " separator
        row_identifier = " - ".join(label_parts)

        # Format data values for this row
        for data_col in data_cols:
            if data_col["index"] < len(row):
                value = row[data_col["index"]]
                if value is not None and str(value).strip():
                    header = data_col["header"]
                    formatted_value = self._format_value(value, data_col.get("type", "unknown"))
                    lines.append(f"- {row_identifier} - {header} = {formatted_value}")

        return lines

    def _format_hierarchical_row(self, row: List, columns: List[Dict],
                                category_col: Dict, subcategory_col: Optional[Dict],
                                data_cols: List[Dict], current_category: Optional[str]) -> List[str]:
        """Format a row with hierarchical structure."""
        lines = []

        # Get category and subcategory values
        category_value = None
        if category_col and category_col["index"] < len(row):
            val = row[category_col["index"]]
            if val is not None and str(val).strip():
                category_value = str(val).strip()

        subcategory_value = None
        if subcategory_col and subcategory_col["index"] < len(row):
            val = row[subcategory_col["index"]]
            if val is not None and str(val).strip():
                subcategory_value = str(val).strip()

        # Use current category if no category in this row
        effective_category = category_value if category_value else current_category

        # Build row identifier
        if effective_category and subcategory_value:
            row_identifier = f"{effective_category}: {subcategory_value}"
        elif effective_category:
            row_identifier = effective_category
        else:
            # No clear category structure
            return self._format_flat_row(row, columns)

        # Format data values
        for data_col in data_cols:
            if data_col["index"] < len(row):
                value = row[data_col["index"]]
                if value is not None and str(value).strip():
                    header = data_col["header"]
                    formatted_value = self._format_value(value, data_col.get("type", "unknown"))
                    lines.append(f"- {row_identifier} - {header} = {formatted_value}")

        return lines

    def _format_flat_row(self, row: List, columns: List[Dict]) -> List[str]:
        """Format row as flat structure."""
        lines = []

        # If we have multiple columns and first looks like identifier
        if len(columns) >= 2 and columns[0].get("role") in ["category", "text"]:
            first_col = row[0] if len(row) > 0 else None
            if first_col is not None and str(first_col).strip():
                row_identifier = str(first_col).strip()

                # Format remaining columns
                for col_idx in range(1, len(columns)):
                    if col_idx < len(row):
                        value = row[col_idx]
                        if value is not None and str(value).strip():
                            header = columns[col_idx]["header"]
                            formatted_value = self._format_value(value, columns[col_idx].get("type", "unknown"))
                            lines.append(f"- {row_identifier} - {header} = {formatted_value}")
        else:
            # Simple column by column
            for col_idx, col_info in enumerate(columns):
                if col_idx < len(row):
                    value = row[col_idx]
                    if value is not None and str(value).strip():
                        header = col_info["header"]
                        formatted_value = self._format_value(value, col_info.get("type", "unknown"))
                        lines.append(f"- {header}: {formatted_value}")

        return lines

    def _find_column_by_role(self, columns: List[Dict], role: str) -> Optional[Dict]:
        """Find first column with specified role."""
        for col in columns:
            if col.get("role") == role:
                return col
        return None

    def _extract_group_key(self, line: str) -> str:
        """Extract a group key from a formatted line for section organization."""
        if not line.startswith("- "):
            return "misc"

        # Remove the "- " prefix and extract the grouping pattern
        content = line[2:]

        # Split on " = " to get the left part (identifier)
        if " = " in content:
            identifier = content.split(" = ")[0]
        elif ": " in content:
            identifier = content.split(": ")[0]
        else:
            identifier = content

        # Extract meaningful grouping from patterns like:
        # "Employee Only - VSP - Premium - Monthly Premium (Funding Rates)"
        # "Employee Only - UHC - PPO - Monthly Pre Tax EE - Hourly"

        parts = identifier.split(" - ")

        if len(parts) >= 3:
            # Look for patterns like: Coverage - Provider - Plan
            # Employee Only - VSP - Premium
            # Employee Only - UHC - PPO
            coverage = parts[0]  # Employee Only
            provider = parts[1]  # VSP, UHC
            plan = parts[2]      # Premium, Buy-Up, Base, PPO

            # Create group key
            group_key = f"{coverage} - {provider} - {plan}"
            return group_key

        elif len(parts) >= 2:
            # Simpler pattern: Coverage - Provider
            return f"{parts[0]} - {parts[1]}"

        else:
            # Single part or unknown pattern
            return parts[0] if parts else "misc"

    def _add_section_headers(self, all_lines: List[Tuple[str, str]]) -> List[str]:
        """Chunk data into 21-row sections with H2 headers. Each section is self-contained."""
        content = []

        # If no lines, return empty
        if not all_lines:
            return content

        # Group lines by their natural grouping first
        grouped_sections = []
        current_section = []
        current_group = None

        for line, group_key in all_lines:
            # Check if we're starting a new natural group
            if group_key != current_group and current_group is not None:
                # Save the current section if it has content
                if current_section:
                    grouped_sections.append((current_group, current_section))
                    current_section = []

            # Update current group
            current_group = group_key
            current_section.append(line)

        # Don't forget the last section
        if current_section:
            grouped_sections.append((current_group, current_section))

        # Now chunk each grouped section considering total lines = 21 max
        for group_key, lines in grouped_sections:
            # Each chunk structure (except first):
            # Line 1: (blank separator from previous chunk)
            # Line 2: ## HEADER
            # Line 3: (blank)
            # Lines 4-21: up to 18 data lines
            # Total: 21 lines maximum

            # For the very first chunk in the document:
            # Line 1: ## HEADER
            # Line 2: (blank)
            # Lines 3-21: up to 19 data lines

            max_data_lines_first = 19 if not content else 18  # First chunk ever gets 19, others get 18
            max_data_lines_per_chunk = 18  # Standard chunks with separator

            # Process lines in chunks
            chunk_start = 0
            chunk_num = 1

            # Calculate total chunks properly
            remaining_lines = len(lines)
            total_chunks = 0
            temp_start = 0
            while temp_start < len(lines):
                if temp_start == 0 and not content:
                    # Very first chunk gets 19 lines
                    temp_start += 19
                else:
                    # Other chunks get 18 lines
                    temp_start += 18
                total_chunks += 1

            while chunk_start < len(lines):
                # Determine chunk size based on position
                if chunk_start == 0 and not content:
                    # Very first chunk in document
                    chunk_size = min(19, len(lines) - chunk_start)
                else:
                    # All other chunks
                    chunk_size = min(18, len(lines) - chunk_start)

                chunk = lines[chunk_start:chunk_start + chunk_size]

                # Generate header for this chunk
                header_key = group_key if group_key and group_key != "misc" else "Data Section"
                display_header = self._clean_group_header(header_key)

                # Add chunk number if this group spans multiple chunks
                if total_chunks > 1:
                    display_header = f"{display_header} (Part {chunk_num})"

                # Skip headers - just add data lines directly
                # Add the data lines
                content.extend(chunk)

                # Move to next chunk
                chunk_start += chunk_size
                chunk_num += 1

        return content


    def _clean_group_header(self, group_key: str) -> str:
        """Clean up group key for use as a header."""
        # Remove common prefixes that are repetitive
        if group_key.startswith("Employee Only - "):
            cleaned = group_key[16:]  # Remove "Employee Only - "
        else:
            cleaned = group_key

        # Capitalize properly
        parts = cleaned.split(" - ")
        capitalized_parts = []

        for part in parts:
            # Handle special cases
            if part.upper() in ['VSP', 'UHC', 'PPO', 'HMO', 'COBRA']:
                capitalized_parts.append(part.upper())
            elif part.lower() in ['buy-up', 'buy up']:
                capitalized_parts.append('Buy-Up')
            else:
                capitalized_parts.append(part.title())

        return " - ".join(capitalized_parts)

    def _format_value(self, value, column_type: str) -> str:
        """Format values based on their detected type."""
        if value is None:
            return ""

        value_str = str(value).strip()
        if not value_str:
            return ""

        # Handle different column types
        if column_type == "currency" or self._looks_like_currency(value_str):
            return self._format_currency(value_str)
        elif column_type == "percentage" or self._looks_like_percentage(value_str):
            return self._format_percentage(value_str)
        elif column_type == "numeric" or self._is_numeric_value(value_str):
            return self._format_numeric(value_str)
        else:
            return value_str

    def _looks_like_currency(self, value_str: str) -> bool:
        """Check if value looks like currency."""
        # Only treat as currency if it has an actual $ symbol
        return '$' in value_str

    def _looks_like_percentage(self, value_str: str) -> bool:
        """Check if value looks like percentage."""
        # Only treat as percentage if it has an actual % symbol
        return '%' in value_str

    def _is_numeric_value(self, value_str: str) -> bool:
        """Check if value is numeric."""
        cleaned = value_str.replace(',', '').replace('$', '').replace('%', '').replace('(', '').replace(')', '')
        try:
            float(cleaned)
            return True
        except (ValueError, TypeError):
            return False

    def _is_large_number(self, value_str: str) -> bool:
        """Check if number is large enough to likely be currency."""
        try:
            val = float(value_str)
            return val >= 10  # Numbers 10+ likely currency
        except (ValueError, TypeError):
            return False

    def _format_currency(self, value_str: str) -> str:
        """Format currency values to 2 decimal places."""
        # Remove existing formatting
        cleaned = value_str.replace('$', '').replace(',', '').strip()

        try:
            amount = float(cleaned)
            # Format with 2 decimal places and add commas for thousands
            if amount >= 0:
                return f"${amount:,.2f}"
            else:
                return f"-${abs(amount):,.2f}"
        except (ValueError, TypeError):
            return value_str

    def _format_percentage(self, value_str: str) -> str:
        """Format percentage values."""
        if '%' in value_str:
            # Already has %, just clean up the number part
            cleaned = value_str.replace('%', '').strip()
            try:
                val = float(cleaned)
                return f"{val:.2f}%"
            except (ValueError, TypeError):
                return value_str
        else:
            # Decimal form, convert to percentage
            try:
                val = float(value_str)
                return f"{val * 100:.2f}%"
            except (ValueError, TypeError):
                return value_str

    def _format_numeric(self, value_str: str) -> str:
        """Format numeric values preserving original precision."""
        cleaned = value_str.replace(',', '').strip()
        try:
            val = float(cleaned)
            # If it's a whole number, show as integer
            if val.is_integer():
                return f"{int(val):,}"
            else:
                # Preserve decimals as-is, just add thousand separators
                # Count decimal places in original
                if '.' in cleaned:
                    decimal_places = len(cleaned.split('.')[-1])
                    return f"{val:,.{decimal_places}f}"
                else:
                    # No decimal in original, but float has decimals
                    return f"{val:,g}"  # Use general format (no trailing zeros)
        except (ValueError, TypeError):
            return value_str

    def _calculate_formatting_confidence(self, content: str, structure: Dict) -> float:
        """Calculate confidence in formatting quality."""
        confidence = 0.5

        # Check content length
        if len(content) > 100:
            confidence += 0.2

        # Check for structured format
        if " - " in content and " = " in content:
            confidence += 0.2

        # Check for data preservation
        lines = content.split("\n")
        data_lines = [l for l in lines if l.strip().startswith("- ")]
        if data_lines:
            confidence += 0.1

        # Boost for structure awareness
        if structure.get("hierarchical") is not None:
            confidence += 0.1

        return min(1.0, confidence)