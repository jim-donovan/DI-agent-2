"""
Excel Ingestion Agent
Processes Excel files with multiple tabs and outputs structured data
"""

import pandas as pd
import numpy as np
from typing import Dict, Any, List, Tuple, Optional
from pathlib import Path
import time
import openpyxl

from agent_base import BaseAgent, AgentResponse
from logger import ProcessingLogger
from excel_structure_agent import ExcelStructureAgent
from excel_formatting_agent import ExcelFormattingAgent


class ExcelIngestionAgent(BaseAgent):
    """Agent for processing Excel files with multi-tab support."""

    def __init__(self, logger: ProcessingLogger, api_client=None):
        super().__init__("excel_ingestion_agent", logger, api_client)
        self.supported_formats = ['.xlsx', '.xls', '.csv']

        # Initialize sub-agents
        self.structure_agent = ExcelStructureAgent(logger, api_client)
        self.formatting_agent = ExcelFormattingAgent(logger, api_client)

    def process(self, input_data: Any, context: Dict[str, Any] = None) -> AgentResponse:
        """Process Excel file using coordinated sub-agents."""
        start_time = time.time()
        context = context or {}

        try:
            file_path = input_data.get("file_path")
            if not file_path:
                raise ValueError("file_path is required in input_data")

            file_path = Path(file_path)
            if not file_path.exists():
                raise FileNotFoundError(f"Excel file not found: {file_path}")

            if file_path.suffix.lower() not in self.supported_formats:
                raise ValueError(f"Unsupported format: {file_path.suffix}")

            # Handle CSV files separately
            if file_path.suffix.lower() == '.csv':
                return self._process_csv_file(file_path, context, start_time)

            # Open workbook and process each sheet
            wb = openpyxl.load_workbook(file_path, data_only=True)
            document_name = file_path.stem

            all_content = []
            total_sheets = len(wb.sheetnames)
            total_rows = 0
            total_columns = 0
            overall_confidence = 0.0

            # Skip document header - just output data

            # Process each sheet with agent coordination
            for sheet_name in wb.sheetnames:
                worksheet = wb[sheet_name]

                # Check if user provided structure configuration
                user_structure = context.get("user_structure")

                if user_structure:
                    self.logger.log_step(f"📐 Using user-configured structure: {sheet_name}")
                    # Use user-provided structure directly
                    structure = user_structure
                else:
                    self.logger.log_step(f"🔍 Analyzing structure: {sheet_name}")

                    # Step 1: Detect structure using structure agent
                    structure_input = {
                        "worksheet": worksheet,
                        "sheet_name": sheet_name
                    }

                    structure_response = self.structure_agent.process(structure_input, context)

                    if not structure_response.success:
                        self.logger.log_warning(f"Structure detection failed for {sheet_name}: {structure_response.error_message}")
                        all_content.extend([
                            f"## Sheet: {sheet_name}",
                            "- *[Structure detection failed]*",
                            ""
                        ])
                        continue

                    structure = structure_response.content

                self.logger.log_step(f"📝 Formatting data: {sheet_name}")

                # Step 2: Format data using formatting agent
                formatting_input = {
                    "structure": structure,
                    "data_source": worksheet,
                    "document_name": document_name,
                    "sheet_name": sheet_name
                }

                formatting_response = self.formatting_agent.process(formatting_input, context)

                if formatting_response.success:
                    # Add the formatted content directly (no headers to skip anymore)
                    sheet_content = formatting_response.content
                    all_content.append(sheet_content)
                    all_content.append("")

                    # Update stats
                    lines = sheet_content.split('\n')
                    data_lines = [line for line in lines if line.strip().startswith("- ")]
                    total_rows += len(data_lines)
                    total_columns = max(total_columns, structure.get("dimensions", {}).get("columns", 0))

                    overall_confidence += formatting_response.confidence
                else:
                    self.logger.log_warning(f"Formatting failed for {sheet_name}: {formatting_response.error_message}")
                    all_content.extend([
                        "- *[Formatting failed]*",
                        ""
                    ])

            wb.close()

            # Calculate overall confidence
            if total_sheets > 0:
                overall_confidence = overall_confidence / total_sheets
            else:
                overall_confidence = 0.0

            processing_time = time.time() - start_time

            self.add_memory("excel_processed", {
                "file_path": str(file_path),
                "sheets_count": total_sheets,
                "processing_time": processing_time,
                "used_agent_coordination": True
            })

            return AgentResponse(
                success=True,
                content="\n".join(all_content),
                confidence=overall_confidence,
                metadata={
                    "file_path": str(file_path),
                    "sheets_processed": total_sheets,
                    "total_rows": total_rows,
                    "total_columns": total_columns,
                    "processing_method": "agent_coordination"
                },
                processing_time=processing_time
            )

        except Exception as e:
            processing_time = time.time() - start_time
            self.logger.log_error(f"Excel processing failed: {str(e)}")

            return AgentResponse(
                success=False,
                content="",
                confidence=0.0,
                error_message=str(e),
                processing_time=processing_time
            )

    def _process_csv_file(self, file_path: Path, context: Dict[str, Any], start_time: float) -> AgentResponse:
        """Process CSV file using pandas."""
        try:
            from datetime import datetime

            # Read CSV file
            df = pd.read_csv(file_path)
            document_name = file_path.stem

            # Build markdown content
            all_content = []
            all_content.extend([
                f"# {document_name}",
                "",
                f"**Processed:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"**Source:** CSV file",
                "",
                "---",
                ""
            ])

            # Add sheet section
            all_content.append(f"## Data")
            all_content.append("")

            if df.empty:
                all_content.append("- *[No data in this file]*")
            else:
                # Clean the dataframe
                df = df.fillna('')

                # Format as markdown list
                for idx, row in df.iterrows():
                    # Create row identifier
                    row_id = f"Row {idx + 1}"
                    all_content.append(f"- **{row_id}**")

                    # Add each column value
                    for col in df.columns:
                        value = str(row[col]).strip()
                        if value:
                            all_content.append(f"  - **{col}**: {value}")

                    all_content.append("")

            # Combine content
            final_content = '\n'.join(all_content)

            # Calculate confidence
            confidence = 0.9 if not df.empty else 0.5

            processing_time = time.time() - start_time

            self.logger.log_success(f"CSV processing completed: {len(df)} rows")

            return AgentResponse(
                success=True,
                content=final_content,
                confidence=confidence,
                metadata={
                    "rows": len(df),
                    "columns": len(df.columns),
                    "file_type": "csv",
                    "sheets_processed": 1  # CSV files have one "sheet"
                },
                processing_time=processing_time
            )

        except Exception as e:
            processing_time = time.time() - start_time
            self.logger.log_error(f"CSV processing failed: {str(e)}")

            return AgentResponse(
                success=False,
                content="",
                confidence=0.0,
                error_message=str(e),
                processing_time=processing_time
            )

    def get_system_prompt(self) -> str:
        """Get the system prompt for Excel processing."""
        return """You are an Excel processing coordinator that manages specialized sub-agents.

Your responsibilities:
1. Coordinate structure detection and data formatting agents
2. Ensure proper Excel file handling with merged cells
3. Maintain data relationships and context
4. Produce structured markdown output for vector storage

You work with two specialized agents:
- Structure Detection Agent: Analyzes table layout and organization
- Formatting Agent: Creates structured output based on detected structure"""

    # Legacy methods - keeping for backward compatibility
    def _format_as_markdown_lists(self, excel_data: Dict[str, pd.DataFrame],
                                 document_name: str) -> str:
        """Format Excel data as nested markdown lists."""
        from datetime import datetime

        content = [
            f"# {document_name}",
            "",
            f"**Processed:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"**Source:** Excel file with {len(excel_data)} sheet(s)",
            "",
            "---",
            ""
        ]

        for sheet_name, df in excel_data.items():
            content.append(f"## Sheet: {sheet_name}")
            content.append("")

            if df.empty:
                content.append("- *[No data in this sheet]*")
                content.append("")
                continue

            # Clean the dataframe
            df_clean = self._clean_dataframe(df)

            # Convert to structured lists
            list_content = self._dataframe_to_lists(df_clean)
            content.extend(list_content)
            content.append("")

        return "\n".join(content)

    def _format_as_tables(self, excel_data: Dict[str, pd.DataFrame],
                         document_name: str) -> str:
        """Format Excel data as markdown tables."""
        from datetime import datetime

        content = [
            f"# {document_name}",
            "",
            f"**Processed:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"**Source:** Excel file with {len(excel_data)} sheet(s)",
            "",
            "---",
            ""
        ]

        for sheet_name, df in excel_data.items():
            content.append(f"## Sheet: {sheet_name}")
            content.append("")

            if df.empty:
                content.append("*[No data in this sheet]*")
                content.append("")
                continue

            # Clean the dataframe
            df_clean = self._clean_dataframe(df)

            # Convert to markdown table
            table_md = df_clean.to_markdown(index=False, tablefmt='pipe')
            content.append(table_md)
            content.append("")

        return "\n".join(content)

    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and prepare dataframe for output."""
        # Make a copy to avoid modifying original
        df_clean = df.copy()

        # Replace NaN values with empty strings
        df_clean = df_clean.fillna('')

        # Clean column names
        df_clean.columns = [str(col).strip() for col in df_clean.columns]

        # Remove completely empty rows
        df_clean = df_clean.dropna(how='all')

        # Convert all values to strings and strip whitespace
        for col in df_clean.columns:
            df_clean[col] = df_clean[col].astype(str).str.strip()

        # Remove rows where all values are empty strings
        df_clean = df_clean[~(df_clean == '').all(axis=1)]

        return df_clean

    def _dataframe_to_lists(self, df: pd.DataFrame) -> List[str]:
        """Convert dataframe to structured list matching the exact format required."""
        if df.empty:
            return ["- *[No data]*"]

        content = []
        columns = list(df.columns)
        current_category = None

        # Process with hierarchical structure if we have multiple columns
        if len(columns) >= 2:
            for _, row in df.iterrows():
                # Get first column value (primary category like "Kaiser - HMO")
                first_col = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""

                # Get second column value (sub-category like "Employee Only")
                second_col = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ""

                # Update category if first column has a value
                if first_col:
                    current_category = first_col

                # Build the formatted output
                if len(columns) > 2:  # We have data columns beyond the first two
                    if current_category and second_col:
                        # Format: Category: Subcategory - Column = Value
                        row_identifier = f"{current_category}: {second_col}"
                        for col_idx in range(2, len(columns)):
                            value = str(row.iloc[col_idx]).strip()
                            if value and value not in ['', '0', '0.0', 'nan']:
                                col_name = self._clean_column_name(str(columns[col_idx]))
                                content.append(f"- {row_identifier} - {col_name} = {value}")
                    elif current_category:
                        # Only category, no subcategory
                        for col_idx in range(1, len(columns)):
                            value = str(row.iloc[col_idx]).strip()
                            if value and value not in ['', '0', '0.0', 'nan']:
                                col_name = self._clean_column_name(str(columns[col_idx]))
                                content.append(f"- {current_category} - {col_name} = {value}")
                else:
                    # Only 2 columns total
                    if first_col and second_col:
                        content.append(f"- {first_col} = {second_col}")
        else:
            # Single column - just list values
            for _, row in df.iterrows():
                value = str(row.iloc[0]).strip()
                if value and value not in ['', 'nan']:
                    content.append(f"- {value}")

        return content if content else ["- *[No readable data]*"]

    def _is_identifier_column(self, column: pd.Series) -> bool:
        """Check if a column appears to be row identifiers."""
        # Count non-empty values
        non_empty = column.astype(str).str.strip().replace('', np.nan).notna().sum()

        # If most rows have values, likely an identifier column
        if non_empty > len(column) * 0.5:
            # Check if values are mostly non-numeric
            numeric_count = 0
            for val in column:
                val_str = str(val).strip()
                if val_str:
                    try:
                        float(val_str.replace(',', '').replace('$', '').replace('%', ''))
                        numeric_count += 1
                    except (ValueError, TypeError):
                        pass

            # If less than half are pure numbers, likely identifiers
            return numeric_count < non_empty * 0.5

        return False

    def _is_category_header(self, text: str) -> bool:
        """Check if text represents a major category or group header."""
        # Generic check: consider it a category if it's:
        # - Short (less than 50 chars)
        # - Not purely numeric
        # - Not empty
        # - Appears to be a label/header (no sentences)

        if not text or len(text) > 50:
            return False

        # Check if it's purely numeric
        if self._is_pure_numeric(text):
            return False

        # Check if it looks like a sentence (has multiple spaces)
        word_count = len(text.split())
        if word_count > 5:  # Likely a sentence, not a category
            return False

        # Check if it has typical data patterns (dates, long numbers, etc.)
        import re
        if re.match(r'^\d{2,4}[-/]\d{1,2}[-/]\d{1,4}$', text):  # Date pattern
            return False
        if re.match(r'^\d{5,}$', text):  # Long number (ID, phone, etc.)
            return False

        # If it passed all checks, consider it a potential category
        return True

    def _clean_column_name(self, col_name: str) -> str:
        """Clean and format column names for display."""
        # Remove common prefixes/suffixes
        col_name = col_name.replace('_', ' ').strip()

        # Handle "Unnamed" columns
        if 'Unnamed' in col_name:
            return 'Value'

        # Capitalize appropriately
        words = col_name.split()
        cleaned_words = []

        for word in words:
            # Keep acronyms uppercase
            if word.isupper() and len(word) <= 4:
                cleaned_words.append(word)
            else:
                cleaned_words.append(word.capitalize())

        return ' '.join(cleaned_words)

    def _process_excel_with_merged_cells(self, file_path: Path, context: Dict[str, Any]) -> str:
        """Process Excel file with proper merged cell handling."""
        from datetime import datetime

        # Open workbook with openpyxl to get merged cell info
        wb = openpyxl.load_workbook(file_path, data_only=True)

        document_name = file_path.stem
        content = [
            f"# {document_name}",
            "",
            f"**Processed:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"**Source:** Excel file with {len(wb.sheetnames)} sheet(s)",
            "",
            "---",
            ""
        ]

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            content.append(f"## Sheet: {sheet_name}")
            content.append("")

            # Process sheet with merged cell awareness
            sheet_content = self._process_sheet_with_merged_cells(ws)
            content.extend(sheet_content)
            content.append("")

        wb.close()
        return "\n".join(content)

    def _process_sheet_with_merged_cells(self, worksheet) -> List[str]:
        """Process a single worksheet with merged cell handling."""
        if worksheet.max_row == 0 or worksheet.max_column == 0:
            return ["- *[No data in this sheet]*"]

        # Build merged cells map
        merged_cells_map = {}
        for merged_range in worksheet.merged_cells.ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            max_row, max_col = merged_range.max_row, merged_range.max_col

            # Get value from top-left cell
            value = worksheet.cell(min_row, min_col).value

            # Map all cells in range to this value
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged_cells_map[(row, col)] = value

        # Read all data including merged cells
        data_rows = []
        for row_idx in range(1, worksheet.max_row + 1):
            row_data = []
            for col_idx in range(1, worksheet.max_column + 1):
                if (row_idx, col_idx) in merged_cells_map:
                    value = merged_cells_map[(row_idx, col_idx)]
                else:
                    value = worksheet.cell(row_idx, col_idx).value
                row_data.append(value)
            data_rows.append(row_data)

        # Detect header rows and data start
        header_indices, data_start = self._detect_header_rows(data_rows)

        # Build column headers from detected header rows
        column_headers = self._build_column_headers(data_rows, header_indices)

        # Format data with proper structure
        return self._format_merged_data(data_rows[data_start:], column_headers)

    def _detect_header_rows(self, data_rows: List[List]) -> Tuple[List[int], int]:
        """Detect which rows are headers vs data."""
        header_indices = []
        data_start = len(data_rows)

        for idx, row in enumerate(data_rows[:min(10, len(data_rows))]):
            non_empty = sum(1 for cell in row if cell is not None and str(cell).strip())
            if non_empty == 0:
                continue

            # Check if row is mostly text (likely header)
            text_count = sum(1 for cell in row
                           if cell is not None and not self._is_pure_numeric(str(cell)))

            if text_count > len(row) * 0.5:
                header_indices.append(idx)
            else:
                data_start = idx
                break

        return header_indices, data_start

    def _is_pure_numeric(self, value: str) -> bool:
        """Check if value is purely numeric."""
        value = str(value).strip()
        test_value = value.replace(',', '').replace('$', '').replace('%', '').replace('(', '').replace(')', '')
        try:
            float(test_value)
            return True
        except (ValueError, TypeError):
            return False

    def _build_column_headers(self, data_rows: List[List], header_indices: List[int]) -> List[str]:
        """Build comprehensive column headers from multi-level headers."""
        if not header_indices:
            num_cols = len(data_rows[0]) if data_rows else 0
            return [f"Column{i+1}" for i in range(num_cols)]

        combined_headers = []
        num_cols = len(data_rows[0]) if data_rows else 0

        for col_idx in range(num_cols):
            header_parts = []

            for header_idx in header_indices:
                if header_idx < len(data_rows):
                    value = data_rows[header_idx][col_idx] if col_idx < len(data_rows[header_idx]) else None
                    if value is not None and str(value).strip():
                        value_str = str(value).strip()
                        # Don't repeat the same value
                        if not header_parts or header_parts[-1] != value_str:
                            header_parts.append(value_str)

            if header_parts:
                combined_headers.append(" - ".join(header_parts))
            else:
                combined_headers.append(f"Column{col_idx+1}")

        return combined_headers

    def _format_merged_data(self, data_rows: List[List], column_headers: List[str]) -> List[str]:
        """Format data rows with full context preservation for vector storage."""
        if not data_rows:
            return ["- *[No data]*"]

        content = []
        current_category = None

        # Process each row
        for row in data_rows:
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            # First column is the primary identifier (e.g., "Kaiser - HMO")
            first_col = str(row[0]).strip() if row[0] is not None else ""

            # Check if we have at least 2 columns for hierarchical structure
            if len(row) > 1 and len(column_headers) > 1:
                # Second column might be sub-category (e.g., "Employee Only")
                second_col = str(row[1]).strip() if row[1] is not None else ""

                # If first column has value, update current category
                if first_col:
                    current_category = first_col

                # Build the row identifier
                if current_category and second_col:
                    # We have both category and sub-category
                    row_identifier = f"{current_category}: {second_col}"

                    # Process remaining columns (data columns)
                    for col_idx in range(2, len(row)):
                        if col_idx < len(column_headers):
                            value = row[col_idx]
                            if value is not None and str(value).strip():
                                header = column_headers[col_idx]
                                # Format: Category: Subcategory - Column Header = Value
                                content.append(f"- {row_identifier} - {header} = {str(value).strip()}")
                elif current_category and not second_col:
                    # Only have category, process all remaining columns
                    for col_idx in range(1, len(row)):
                        if col_idx < len(column_headers):
                            value = row[col_idx]
                            if value is not None and str(value).strip():
                                header = column_headers[col_idx]
                                content.append(f"- {current_category} - {header} = {str(value).strip()}")
                elif first_col:
                    # First column has data, use it as the identifier
                    for col_idx in range(1, len(row)):
                        if col_idx < len(column_headers):
                            value = row[col_idx]
                            if value is not None and str(value).strip():
                                header = column_headers[col_idx]
                                content.append(f"- {first_col} - {header} = {str(value).strip()}")
            else:
                # Single column or no clear structure
                for col_idx, value in enumerate(row):
                    if value is not None and str(value).strip():
                        if col_idx < len(column_headers):
                            header = column_headers[col_idx]
                            content.append(f"- {header}: {str(value).strip()}")

        return content if content else ["- *[No readable data]*"]

    def _is_first_column_labels(self, data_rows: List[List]) -> bool:
        """Check if first column contains row labels."""
        if not data_rows:
            return False

        non_empty_first = sum(1 for row in data_rows
                            if row and row[0] is not None and str(row[0]).strip())

        return non_empty_first > len(data_rows) * 0.5

    def _calculate_content_confidence(self, content: str) -> float:
        """Calculate confidence based on content quality."""
        base_confidence = 0.9

        if len(content) < 100:
            base_confidence -= 0.3
        elif len(content) < 500:
            base_confidence -= 0.1

        if '- **' in content:
            base_confidence += 0.05
        if '## Sheet:' in content:
            base_confidence += 0.05

        has_numbers = any(char.isdigit() for char in content)
        if not has_numbers:
            base_confidence -= 0.2

        return max(0.1, min(1.0, base_confidence))

    def _calculate_excel_confidence(self, excel_data: Dict[str, pd.DataFrame],
                                   content: str) -> float:
        """Calculate confidence score for Excel processing."""
        base_confidence = 0.9

        # Reduce confidence for empty sheets
        total_sheets = len(excel_data)
        empty_sheets = sum(1 for df in excel_data.values() if df.empty)
        if empty_sheets > 0:
            base_confidence -= (empty_sheets / total_sheets) * 0.3

        # Reduce confidence if output is very short
        if len(content) < 100:
            base_confidence -= 0.2

        # Check for data quality indicators
        total_rows = sum(len(df) for df in excel_data.values())
        if total_rows == 0:
            base_confidence = 0.1
        elif total_rows < 5:
            base_confidence -= 0.2

        return max(0.0, min(1.0, base_confidence))



def process_excel_file(file_path: str, output_format: str = "markdown_lists",
                      logger: ProcessingLogger = None) -> Dict[str, Any]:
    """
    Convenience function to process an Excel file.

    Args:
        file_path: Path to Excel file
        output_format: "markdown_lists" or "tables"
        logger: Optional logger instance

    Returns:
        Dict with processing results
    """
    if logger is None:
        logger = ProcessingLogger()

    agent = ExcelIngestionAgent(logger)

    input_data = {
        "file_path": file_path
    }

    context = {
        "output_format": output_format
    }

    response = agent.process(input_data, context)

    return {
        "success": response.success,
        "content": response.content,
        "metadata": response.metadata,
        "confidence": response.confidence,
        "processing_time": response.processing_time,
        "error_message": response.error_message
    }